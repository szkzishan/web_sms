#!/usr/bin/env python
# -*- coding: utf-8 -*-  

# Created by kyh on 2017/11/6
import os

import paramiko
from django.conf import settings
from django.template.loader import render_to_string
import psycopg2
from openpyxl import Workbook
from psycopg2.extensions import ISOLATION_LEVEL_AUTOCOMMIT

from models import *
import codecs
import json
from rest_framework import exceptions

TEMPLATES_PATH = os.path.join(settings.BASE_DIR, "codeeditor", "templates")

from django.forms.models import model_to_dict


# get model log by user
def get_model_log_by_user(user):
    pass


# 获取APP配置
def get_app_config(app_id):
    app = AppModel.objects.get(id=app_id)
    config = {}
    config["tables"] = {}
    for table in app.tables.all():
        config["tables"][table.name] = {}
        for field in table.fields.all():
            config["tables"][table.name][field.name] = model_to_dict(field)
    return config


class CreateProject(object):
    def __init__(self, app):
        self.app = app
        self.app_name = self.app.name
        self.project_name = self.app_name + "_sms"
        self.project_path = os.path.join(settings.LOCAL_PROJECTS_PATH, self.project_name)  # get from settings
        self.project_sms_path = os.path.join(self.project_path, self.project_name)
        self.project_app_path = os.path.join(self.project_path, self.app_name)
        self.tables = self.app.tables.all()
        self.configs = self.get_configs()

    def create_code(self):
        if not os.path.exists(self.project_path):
            os.mkdir(self.project_path)
            # write manage.py
            self.write_file(
                os.path.join(self.project_path, "manage.py"),
                render_to_string("project/manager.py-tpl", {
                    "project_name": self.project_name
                }))
            # write supervisor.ini
            self.write_file(
                os.path.join(self.project_path, "supervisor.ini"),
                render_to_string("project/supervisor.ini-tpl", {
                    "project_name": self.project_name,
                    "server_port": self.app.server_port,
                    "remote_project_path": self.app.server.project_path + "/" + self.project_name,
                }))
        if not os.path.exists(self.project_sms_path):
            os.mkdir(self.project_sms_path)
            # change main to package open("test.txt","w+").close()
            open(os.path.join(self.project_sms_path, "__init__.py"), "w").close()
            # write settings.py
            self.write_file(
                os.path.join(self.project_sms_path, "settings.py"),
                render_to_string("project/project/settings.py-tpl", {
                    "app_name": self.app_name,
                    "project_name": self.project_name,
                    "database_config": None
                }))
            # write urls.py
            self.write_file(
                os.path.join(self.project_sms_path, "urls.py"),
                render_to_string("project/project/urls.py-tpl", {
                    "app_name": self.app_name,
                    "app_des": self.app.des,
                }))
            # write wsgi.py
            self.write_file(
                os.path.join(self.project_sms_path, "wsgi.py"),
                render_to_string("project/project/wsgi.py-tpl", {
                    "project_name": self.project_name,
                }))
        if not os.path.exists(self.project_app_path):
            os.mkdir(self.project_app_path)
            # change app_dir to package
            open(os.path.join(self.project_app_path, "__init__.py"), "w").close()
            # write apps.py
            self.write_file(
                os.path.join(self.project_app_path, "apps.py"),
                render_to_string("project/app/apps.py-tpl", {
                    "app_name": self.app_name,
                }))
            # write urls.py
            self.write_file(
                os.path.join(self.project_app_path, "urls.py"),
                render_to_string("project/app/urls.py-tpl", {

                }))
        self.write_models_file()
        self.write_serializers_file()
        self.write_filters_file()
        self.write_views_file()

    def get_configs(self):
        configs = {}
        config_path = os.path.join(settings.BASE_DIR, "codeeditor", "config")
        for file in os.listdir(config_path):
            configs[os.path.splitext(file)[0]] = json.loads(open(os.path.join(config_path, file)).read())
        return configs

    def write_file(self, file_name, code):
        # 写入文件
        with codecs.open(file_name, 'w', 'utf-8') as file:
            file.write(code)
            file.write("\n")
            file.close()

    def write_models_file(self):
        model_config = self.configs.get("models")
        modeldefault_config = {
          "CharField": "",
          "IntegerField": "0",
          "FloatField": "0.00",
          "TextField": "",
          "DateTimeField": "False",
          "DateField": "False",
          "BooleanField": "False"
        }

        def get_default_config(fieldtype, defaultValue):
            if defaultValue:
                return defaultValue

            return modeldefault_config[fieldtype]

        tables = {}
        # 遍历表
        for table in self.tables:
            tables[table] = []
            # 遍历字段
            for field in table.fields.all():

                field_str = model_config.get(field.genre) % {
                    "field_name": field.name,
                    "field_type": field.genre,
                    "des": field.des,
                    "param": field.param,
                    "related_name": "_".join([table.name, field.name]),
                    "fdefault":get_default_config(field.genre, field.default_value)
                }
                tables[table].append(field_str)
        code = render_to_string("project/app/models.py-tpl", {
            "tables": tables,
        })
        self.write_file(os.path.join(self.project_app_path, "models.py"), code)

    def write_serializers_file(self):
        serializer_config = self.configs.get("serializers")
        relate_fields = {}
        foreign_fields = {}
        table_extra_fields_forgin = {}
        table_extra_fields_relate = {}
        for table in self.tables:
            # 外键字段处理
            foreign_fields[table.name] = []
            table_extra_fields_forgin[table.name] = []
            for field in FieldModel.objects.filter(table=table):
                if field.genre == "ForeignKey":
                    # 外键详情序列化
                    foreign_fields[table.name].append(serializer_config.get("foreign_field_template") % {
                        "field": field.name,
                        "table": field.param
                    })
                    table_extra_fields_forgin[table.name].append(field.name + "_get")
                    # 主表 相关性 字段
                    relate_filde = serializer_config.get("relate_field_template") % {
                        "field": table.name + "_" + field.name,
                        "table": table.name
                    }
                    if relate_fields.has_key(field.param):
                        relate_fields[field.param].append(relate_filde)
                    else:
                        relate_fields[field.param] = [relate_filde, ]
                    if table_extra_fields_relate.has_key(field.param):
                        table_extra_fields_relate[field.param].append(field.name + "_get")
                    else:
                        table_extra_fields_relate[field.param] = [field.name + "_get", ]
        # 相关字段 field 加入 serializer
        # table 基本序列化
        tables_name = [table.name for table in self.tables]
        tables = {}
        for table in self.tables:
            tables[table.name] = {
                "foreign_fields": foreign_fields.get(table.name),
                "relate_fields": relate_fields.get(table.name),
                "extra_fields": str(
                    table_extra_fields_forgin.get(table.name, []) + table_extra_fields_relate.get(table.name, [])),
            }

        code = render_to_string("project/app/serializers.py-tpl", {
            "tables": tables,
            "tables_name": tables_name
        })
        self.write_file(os.path.join(self.project_app_path, "serializers.py"), code)

    def write_views_file(self):
        tables = {}
        for table in self.tables:
            tables[table.name] = {
                "search_fields": [],
            }
            for field in FieldModel.objects.filter(table=table):
                if field.is_search:
                    tables[table.name]["search_fields"].append(field.name)
        code = render_to_string("project/app/views.py-tpl", {"tables": tables})
        self.write_file(os.path.join(self.project_app_path, "views.py"), code)

    def write_filters_file(self):
        tables = {}
        for table in self.tables:
            tables[table.name] = {
                "relate_fields": [],
                "fields": {"id": "__all__"}
            }
            for field in FieldModel.objects.filter(table=table):
                if field.is_filter:
                    if field.genre == "ForeignKey":
                        tables[table.name]["relate_fields"].append(
                            "%(field_name)s = filterabc.RelatedFilter(%(field_param)sBaseFilter, name='%(field_name)s', queryset=%(field_param)sModel.objects.all())" % {
                                "field_name":field.name,
                                "field_param":field.param,
                            }
                        )
                    else:
                        tables[table.name]["fields"][field.name] = "__all__"

        code = render_to_string("project/app/filters.py-tpl", {"tables": tables})
        self.write_file(os.path.join(self.project_app_path, "filters.py"), code)


class DeployProject(CreateProject):
    def __init__(self, app):
        super(DeployProject, self).__init__(app)
        self.ssh = self._get_ssh()
        self.sftp = self._get_sftp()
        self.remote_home_path = self._get_home_path()
        self.remote_project_path = self._get_remote_path()
        self.bin = "miniconda2/bin/"
        self.manage_command = " ".join([self.bin + "python", self.remote_project_path + "/" + "manage.py"])

    def _get_ssh(self):
        print "_get_ssh"
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(self.app.server.host_ip, self.app.server.host_port, self.app.server.host_user,
                    self.app.server.host_password)
        return ssh

    def _get_sftp(self):
        print "_get_sftp"
        paramiko.SFTPClient.from_transport(self.ssh.get_transport())
        sftp = self.ssh.open_sftp()
        return sftp

    def _get_home_path(self):
        print "_get_home_path"
        _, stdout, stderr = self.ssh.exec_command("cd ~;pwd")
        stdout = stdout.readlines()
        if stdout:
            return stdout[0].strip()

    def _check_dir(self, sftp, path):
        print "_check_dir"
        try:
            sftp.stat(path)
            return True
        except Exception, e:
            return False

    def _info_format(self, exec_status, msg, info):
        print "_info_format"
        return {
            "exec_status": exec_status,
            "msg": msg,
            "info": info
        }

    def _exec_result(self, stdout, stderr, command=u"command"):
        print "_exec_result"
        if stderr:
            return self._info_format(False, " ".join([command, "failure"]), "".join(stderr))
        else:
            return self._info_format(True, " ".join([command, "success"]), "".join(stdout))

    def _get_remote_path(self):
        print "_get_remote_path"
        # 创建并返回 远程工程路径
        project_path = "/".join([self.app.server.project_path, self.app.name + "_sms"])
        self.sftp.chdir(self.app.server.project_path)
        if not self._check_dir(self.sftp, self.app.name + "_sms"):
            self.sftp.mkdir(self.app.name + "_sms")
        return project_path

    def _push_dir(self, local_path, remote_path):
        print "_push_dir"
        self.sftp.chdir(remote_path)
        dirs = []
        for file in os.listdir(local_path):
            file_path = os.path.join(local_path, file)
            if os.path.isfile(file_path):
                self.sftp.put(file_path, file)
            else:
                if not self._check_dir(self.sftp, file):
                    self.sftp.mkdir(file)
                dirs.append(file)
        for dir_name in dirs:
            remote_path_temp = remote_path + "/" + dir_name
            file_path = os.path.join(local_path, dir_name)
            self._push_dir(file_path, remote_path_temp)

    def _create_database(self):
        print "_create_database"
        conn = psycopg2.connect(user="fenghuang", password="fenghuang", host="118.190.68.206", port="5432")
        conn.set_isolation_level(ISOLATION_LEVEL_AUTOCOMMIT)
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) = 0 FROM pg_catalog.pg_database WHERE datname = '%s'" % self.project_name.lower())
        not_exists_row = cur.fetchone()
        not_exists = not_exists_row[0]
        if not_exists:
            cur.execute("CREATE DATABASE %s ENCODING 'UTF8'" % self.project_name.lower())
        cur.close()
        conn.close()

    # 检测环境
    def check_env(self):
        print "check_env"
        self.sftp.chdir(self.remote_home_path)
        if self._check_dir(self.sftp, self.bin):
            return self._info_format(True, u"服务器环境已经存在", None)
        else:
            return self.install_env()

    # 安装环境
    def install_env(self):
        print "install_env"
        env_name = "miniconda2.tar.gz"
        env_path = settings.EVN_PATH
        evn_file = os.path.join(env_path, env_name)
        self.sftp.put(evn_file, self.remote_home_path + "/" + env_name)
        _, stdout, stderr = self.ssh.exec_command("tar -zxvf %s" % env_name)
        return self._exec_result(stdout.readlines(), stderr.readlines(), command=u"evn install")

    # 源码创建件 并传送
    def push_project(self):
        print "push_project"
        self.create_code()
        self.sftp.chdir(self.remote_project_path)
        self._push_dir(self.project_path, self.remote_project_path)
        return self._info_format(True, u"源码创建成功", None)

    # 数据库构建
    def project_migrate(self):
        print "project_migrate"
        # exec makemigrations
        self._create_database()
        _, stdout, stderr = self.ssh.exec_command("%s makemigrations %s" % (self.manage_command, self.app.name))
        self._exec_result(stdout.readlines(), stderr.readlines(), "makemigrations build")
        _, stdout, stderr = self.ssh.exec_command("%s migrate" % self.manage_command)
        return self._exec_result(stdout.readlines(), stderr.readlines(), "database build")

    # 远程启动
    def project_runserver(self):
        print "project_runserver"
        _, stdout, stderr = self.ssh.exec_command("ps -ef | grep %s | awk '{print $2}' | xargs kill -9" % self.app.name)
        self._exec_result(stdout.readlines(), stderr.readlines(), "clean runserver")

        _, stdout, stderr = self.ssh.exec_command(
            "nohup %s runserver 0.0.0.0:%s > %s &" % (
                self.manage_command, self.app.server_port, self.remote_project_path + "/" + "server.log"))
        return self._exec_result(stdout.readlines(), stderr.readlines(), "project runserver")

    # 自动联调
    # 系统档案准备

    # 关闭所有链接
    def clsoe(self):
        print "clsoe"
        self.sftp.close()
        self.ssh.close()


def deploy_project(app_name):
    project_deploy = DeployProject(app_name)
    project_deploy.clsoe()


class ManagerProject(object):
    def __init__(self, user):
        self.user = user
        self.projects = user.user_apps.all()

    def status(self):
        pass

    def start(self):
        pass

    def close(self):
        pass

    def restart(self):
        pass

    def delete(self):
        pass


# 发送验证短信
import requests

HEADER = {
    "Authorization": "APPCODE 34b9048d5ff04aabbdf5067f2315f442"
}

'''
495927（复旦求是注册验证码），此验证码只用于注册你的账号，提供给他人可能导致账号被盗，请勿转发。
'''


def create_sms_template():
    querys = {
        "content": "[code]（复旦求是注册验证码），此验证码只用于注册你的账号，提供给他人可能导致账号被盗，请勿转发。",
        "notiPhone": "18364218861",
        "title": "复旦求是",
    }
    url = "http://ali-sms.showapi.com/createTemplate"
    response = requests.get(url, params=querys, headers=HEADER)
    print response.text


'''
"tNum":"T170317001798"
'''


def send_sms(content, mobile, tNum):
    querys = {
        "content": content,
        "mobile": mobile,
        "tNum": tNum,
    }
    url = "http://ali-sms.showapi.com/sendSms"
    response = requests.get(url, params=querys, headers=HEADER)
    print response.text


def create_database():
    host = ""
    port = ""
    pass


# database excel doc
class AppDoc(object):
    '''
    app_id
    '''

    def __init__(self, app_id):
        self.app = self._get_app(app_id)

    def _get_app(self, app_id):
        return AppModel.objects.get(id=app_id)

    def _get_tables(self):
        return self.app.tables.all()

    def _get_field(self, table):
        return table.fields.all()


from openpyxl.reader.excel import load_workbook


class AppExcelDoc(AppDoc):
    '''

    '''

    def _create_excel(self):
        self.file_path = os.path.join(settings.EXCEL_DOC_PATH, self.app.name + ".xlsx")
        if os.path.exists(self.file_path):
            os.remove(self.file_path)
        else:
            wb = Workbook()
            self.app_excel = wb

    def _write_excel(self):
        wb = self.app_excel
        for table in self._get_tables():
            ws = wb.create_sheet(table.name)
            # ws.row_dimensions[1].height = 70
            for colt in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws.column_dimensions[colt].width = 30
            fields = self._get_field(table)
            ws_header = [u"字段", u"字段名称", u"字段类型", u"字段简述", u"备注"]
            ws.append(ws_header)
            for field in fields:
                col = [
                    field.name,
                    field.des,
                    field.genre,
                    field.des,
                    "",
                ]
                ws.append(col)
        wb.save(self.file_path)

    def make_doc(self):
        self._create_excel()
        self._write_excel()


# json api
import codecs


def json_api():
    print "abxc"
    methods = {
        "list": "get",
        "read": "get",
        "delete": "delete",
        "create": "post",
        "update": "put",
        "partial_update": "patch",
    }
    api = []
    app = AppModel.objects.get(id=347)
    tables = app.tables.all()
    for table in tables:

        for k, v in methods.items():
            json_doc = {}
            json_doc["name"] = "_".join([table.name, k])
            json_doc["desc"] = table.des
            json_doc["method"] = v
            json_doc["endpoint"] = "/".join([app.name, table.name]) + "/"
            if k in ["read", "delete", "update", "partial_update"]:
                json_doc["endpoint"] = "/".join([app.name, table.name, "{id}"]) + "/"
            params = {}
            if v == "post":
                fields = table.fields.all()
                for field in fields:
                    params[field.name] = field.des
            json_doc["params"] = params

            api.append(json_doc)
    print api
    with codecs.open("d:/api.json", "w", 'utf-8') as tt:
        tt.write(json.dumps(api, ensure_ascii=False, indent=2))
        tt.close()


config_data = {
    "table_name": {
        "field_name": "bj"
    }
}
